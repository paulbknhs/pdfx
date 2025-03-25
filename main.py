import os
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from PyPDF2 import PdfReader
import openpyxl
from openpyxl import Workbook

EXCEL_FILE = "Formulardaten.xlsx"
PDF_DIR = "."

class PDFHandler(FileSystemEventHandler):
    def __init__(self):
        super().__init__()
        self.headers_created = False
        self.init_excel()

    def init_excel(self):
        """Initialize the Excel file if it does not exist."""
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Formulardaten"
            wb.save(EXCEL_FILE)
            wb.close()
            print("Neue Excel-Datei erstellt")
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            if ws.max_row > 0:
                self.headers_created = True
            wb.close()

    def on_created(self, event):
        """Triggered when a new PDF is added to the folder."""
        if event.is_directory or not event.src_path.lower().endswith(".pdf"):
            return

        time.sleep(0.5)
        self.process_pdf(event.src_path)

    def save_workbook_with_retry(self, wb, filename, retries=5, delay=1):
        """Tries to save the Excel file with retries in case of a file lock."""
        for attempt in range(retries):
            try:
                wb.save(filename)
                wb.close()
                return
            except PermissionError:
                print(f"Excel-Datei ist gesperrt, erneuter Versuch {attempt+1}/{retries} in {delay} Sekunden...")
                time.sleep(delay)
        print("Fehler: Konnte die Excel-Datei nicht speichern.")

    def delete_pdf_with_retry(self, pdf_path, delay=1):
        """Tries to delete the PDF file with retries in case it's still locked."""
        while (1):
            try:
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                    print("PDF erfolgreich gelöscht")
                    return
            except PermissionError:
                print(f"PDF-Datei ist gesperrt, erneuter Löschversuch in {delay} Sekunden...")
                time.sleep(delay)
        print(f"Fehler: Konnte PDF {pdf_path} nicht löschen.")

    def process_pdf(self, pdf_path):
        """Extracts form fields from a PDF and saves them to the Excel file."""
        try:
            print(f"\nVerarbeite Datei: {os.path.basename(pdf_path)}")

            # Formularfelder extrahieren
            fields = self.extract_form_fields(pdf_path)
            if not fields:
                print("Keine Formularfelder gefunden")
                return

            print("Gefundene Felder:", fields)

            # Load Excel workbook
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

            # Add header row if it doesn't exist
            if not self.headers_created:
                ws.append(list(fields.keys()))
                self.headers_created = True
                print("Header geschrieben")

            # Ensure fields match the existing header order
            try:
                ordered_data = [str(fields.get(header.value, "")) for header in ws[1]]
                ws.append(ordered_data)
                print("Datenzeile hinzugefügt:", ordered_data)
            except KeyError as e:
                print(f"Fehler: Feld {e} nicht in Excel-Header gefunden")

            # Save workbook with retry
            self.save_workbook_with_retry(wb, EXCEL_FILE)

            # Ensure the PDF file is closed before deletion
            time.sleep(0.5)  # Short delay to ensure file is fully released

            # Attempt to delete the PDF with retries
            self.delete_pdf_with_retry(pdf_path)

        except Exception as e:
            print(f"Kritischer Fehler: {str(e)}")

    def extract_form_fields(self, pdf_path):
        """Extracts form fields from a PDF file."""
        try:
            with open(pdf_path, "rb") as f:
                reader = PdfReader(f)
                fields = reader.get_form_text_fields()
                if not fields:
                    return None
                
                return {name: fields[name] or "" for name in fields}
        except Exception as e:
            print(f"PDF-Lesefehler: {str(e)}")
            return None

if __name__ == "__main__":
    event_handler = PDFHandler()
    observer = Observer()
    observer.schedule(event_handler, PDF_DIR)

    print(f"Überwache Ordner '{os.path.abspath(PDF_DIR)}'...")
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    
    observer.join()
